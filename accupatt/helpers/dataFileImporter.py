import io
import os

import accupatt.config as cfg
import numpy as np
import openpyxl
import pandas as pd
from accupatt.helpers.dBBridge import save_to_db
from accupatt.models.appInfo import AppInfo, Nozzle
from accupatt.models.passData import Pass
from accupatt.models.seriesData import SeriesData
from accupatt.models.sprayCard import SprayCard
from openpyxl_image_loader import SheetImageLoader


def convert_xlsx_to_db(file, s: SeriesData = None) -> str:
    # Only load in series data from xlsx if no series passed in
    if s == None:
        s = load_from_accupatt_1_file(file)
    #Write to DB (same dir as original xlsx)
    file_db = os.path.splitext(file)[0]+'.db'
    save_to_db(file=file_db, s=s)
    #Append Images
    wb = openpyxl.load_workbook(file)
    if not 'Card Data' in wb.sheetnames:
        return s
    sh = wb['Card Data']
    on_pass = int(sh['B2'].value)
    # Loop over declard cards and save images to db if has_image
    p: Pass = s.passes[on_pass-1]
    c: SprayCard
    for c in p.spray_cards:
        c.filepath = file_db
        if c.has_image:
            # Get the image from applicable sheet
            image_loader = SheetImageLoader(wb[c.name])
            image = image_loader.get('A1')
            # Conver to bytestream
            stream = io.BytesIO()
            image.save(stream, format="PNG")
            # Save it to the database
            c.save_image_to_file(stream.getvalue())
            # Reclaim resources
            stream.close()
            
    return file_db

def load_from_accupatt_1_file(file) -> SeriesData:
    #indicator for metric
    isMetric = False

    #Load entire WB into dict of sheets
    df_map = pd.read_excel(file, sheet_name=None, header=None)

    #initialize SeriesData object to store all info
    s = SeriesData()
    i = s.info
    #Pull data from Fly-In Tab
    df = df_map['Fly-In Data'].fillna('')
    i.flyin_name = df.iat[0,0]
    i.flyin_location = df.iat[1,0]
    i.flyin_date = df.iat[2,0]
    i.flyin_analyst = df.iat[3,0]

    #Pull data from AppInfo Tab
    df = df_map['Aircraft Data'].fillna('')
    i.pilot = df.iat[0,1]
    i.business = df.iat[1,1]
    i.phone = df.iat[2,1]
    i.street = df.iat[3,1]
    i.city = df.iat[4,1]
    i.state = df.iat[5,1]
    i.regnum = df.iat[6,1]
    i.series = df.iat[7,1]
    i.make = df.iat[8,1]
    i.model = df.iat[9,1]
    if((noz_type_1 := df.iat[10,1]) != ''):
        i.nozzles.append(Nozzle(id=1,
                              type=noz_type_1,
                              size=df.iat[12,1],
                              deflection=df.iat[13,1],
                              quantity=df.iat[11,1]))
    if((noz_type_2 := df.iat[14,1]) != ''):
        i.nozzles.append(Nozzle(id=2,
                              type=noz_type_2,
                              size=df.iat[16,1],
                              deflection=df.iat[17,1],
                              quantity=df.iat[15,1]))
    i.set_pressure(df.iat[18,1])
    i.set_rate(df.iat[19,1])
    i.set_swath(df.iat[20,1])
    i.set_swath_adjusted(df.iat[20,1]) # Just in case it isn't set below
    i.set_wingspan(df.iat[27,1])
    i.set_boom_width(df.iat[28,1])
    i.set_boom_drop(df.iat[30,1])
    i.set_nozzle_spacing(df.iat[31,1])
    i.winglets = df.iat[32,1]
    i.notes_setup = df.iat[33,1]
    # Col 2 if available
    if df.shape[1] > 2:
        i.email = df.iat[0,2]
        i.zip = str(df.iat[5,2])
        i.set_swath_adjusted(df.iat[20,2])
    # Set units for series/passes based on 'metric' identifier
    isMetric = (df.iat[35,1] == 'TRUE')
    s.info.swath_units = cfg.UNIT_M if isMetric else cfg.UNIT_FT
    s.info.rate_units = cfg.UNIT_LPHA if isMetric else cfg.UNIT_GPA
    s.info.pressure_units = cfg.UNIT_BAR if isMetric else cfg.UNIT_PSI
    s.info.wingspan_units = cfg.UNIT_M if isMetric else cfg.UNIT_FT
    s.info.boom_width_units = cfg.UNIT_M if isMetric else cfg.UNIT_FT
    s.info.boom_drop_units = cfg.UNIT_CM if isMetric else cfg.UNIT_IN
    s.info.nozzle_spacing_units = cfg.UNIT_CM if isMetric else cfg.UNIT_IN
        
    #Pull data from Series Data tab
    df = df_map['Series Data'].fillna('')
    df.columns = df.iloc[0]
    #Clear any stored individual passes
    s.passes = []
    #Search for any active passes and create entries in seriesData.passes dict
    for column in df.columns[1:]:
        if not str(df.at[1,column]) == '':
            p = Pass(number = df.columns.get_loc(column))
            p.set_ground_speed(df.at[1,column], units=cfg.UNIT_KPH if isMetric else cfg.UNIT_MPH)
            p.set_spray_height(df.at[2,column], units=cfg.UNIT_M if isMetric else cfg.UNIT_FT)
            p.set_pass_heading(df.at[3,column])
            p.set_wind_direction(df.at[4,column])
            p.set_wind_speed(df.at[5,column], units=cfg.UNIT_KPH if isMetric else cfg.UNIT_MPH)
            p.set_temperature(df.at[6,'Pass 1'], units=cfg.UNIT_DEG_C if isMetric else cfg.UNIT_DEG_F)
            p.set_humidity(df.at[7,'Pass 1'])
            p.data_loc_units = cfg.UNIT_M if isMetric else cfg.UNIT_FT
            s.passes.append(p)

    #Pull data from Pattern Data tab
    df: pd.DataFrame = df_map['Pattern Data'].fillna(np.nan)
    #if df.shape[1] < 13: df['nan'] = np.nan
    #Make new empty dataframe for info
    #df_info = pd.DataFrame({'Pass 1':[], 'Pass 2':[], 'Pass 3':[], 'Pass 4':[], 'Pass 5':[], 'Pass 6':[]})
    cols = ['loc', 'Pass 1', 'Pass 2', 'Pass 3', 'Pass 4', 'Pass 5', 'Pass 6']
    trims = df.iloc[0:3,[2, 4, 6, 8, 10, 12]].reset_index(drop=True).astype(float)
    trims.columns = cols[1:]
    params = df.iloc[3:4,[1, 3, 5, 7, 9, 11]].reset_index(drop=True)
    params.append(df.iloc[4:4,[2, 4, 6, 8, 10, 12]].reset_index(drop=True))
    params.columns = cols[1:]
    df_em = df.iloc[5:,[0,2,4,6,8,10,12]].reset_index(drop=True).astype(float)
    df_em.columns = cols
    df_ex = df.iloc[5:,[0,1,3,5,7,9,11]].reset_index(drop=True).astype(float)
    df_ex.columns = cols

    #Pull patterns and place them into seriesData.passes list by name (created above)
    p: Pass
    for p in s.passes:
        p.trim_l = 0 if trims.at[0,p.name] == np.nan else int(trims.at[0,p.name])
        p.trim_r = 0 if trims.at[1,p.name] == np.nan else int(trims.at[1,p.name])
        p.trim_v = 0 if trims.at[2,p.name] == np.nan else trims.at[2,p.name]
        # TODO Integration Time is params row 0, must convert to int
        # TODO Ex/Em Wavs are params rows 1, 2 respective, must strip string
        p.data = df_em[['loc',p.name]]
        p.data_ex = df_ex[['loc',p.name]]                

    #Create SprayCards if applicable
    if 'Card Data' in df_map.keys():
        wb = openpyxl.load_workbook(file, read_only=True)
        sh = wb['Card Data']
        spacing = sh['B1'].value
        on_pass = int(sh['B2'].value)
        spread_method_text = sh['B3'].value
        spread_method = cfg.SPREAD_METHOD_ADAPTIVE
        if spread_method_text == cfg.SPREAD_METHOD_DIRECT_STRING:
            spread_method = cfg.SPREAD_METHOD_DIRECT
        if spread_method_text == cfg.SPREAD_METHOD_NONE_STRING:
            spread_method = cfg.SPREAD_METHOD_NONE
        spread_a = sh['B4'].value
        spread_b = sh['B5'].value
        spread_c = sh['B6'].value
        #Card Pass
        p: Pass = s.passes[on_pass-1]
        for col in range(5,14):
            if not sh.cell(row=1,column=col).value:
                continue
            c = SprayCard(name=sh.cell(row=1,column=col).value)
            c.filepath = file
            c.location = (col-9)*spacing
            c.location_units = cfg.UNIT_FT
            c.has_image = 1 if sh.cell(row=2,column=col).value else 0
            c.include_in_composite = 1 if sh.cell(row=3,column=col).value else 0
            if c.has_image:
                c.threshold_grayscale = sh.cell(row=4,column=col).value
            c.threshold_type = cfg.THRESHOLD_TYPE_GRAYSCALE
            c.spread_method = spread_method
            c.spread_factor_a = spread_a
            c.spread_factor_b = spread_b
            c.spread_factor_c = spread_c
            p.spray_cards.append(c)
            
    return s
